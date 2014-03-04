import os
import math
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook#for writing to excel file

#Limbs
LH = 0
RH = 1
HEAD_PPT_POS = 2
HEAD_PPT_EULER = 3
LH_PPT_EULER = 4
RH_PPT_EULER = 5
DATE = 6
TIME = 7

YAW = 0
PITCH = 1
ROLL = 2

X_POS = 0
Y_POS = 1
Z_POS = 2

TARGET_TIME = 0
FIRE_TIME = 1
INTERVAL = 120
#LIMB_NAMES = ['LeftHand', 'RightHand', 'HeadPosition', 'HeadEuler', 'LeftHandEuler', 'RightHandEuler']
NUM_MIN = 9
ORIENTATIONS = ['', 'YAW' ,'PITCH', 'ROLL', 'X_POS','Y_POS','Z_POS']#0 means orientation not applicable
dirList = os.listdir(os.path.curdir)
dirPre = ""
#FUNCTIONS
def dist(p1, p2, orientation):
    if orientation > 3:
        return abs(p2[orientation] - p1[orientation])
    else:
        distanceCalculated = abs(p2[orientation] - p1[orientation])
        print "points"
        print p2
        print p1
        if distanceCalculated > 180:
            return 360 - distanceCalculated
        else:
            return distanceCalculated
#uses a different distance function to calculate changes in position for the intersense

#outputs a graph to the file with the given limb and orientation movement graph
def makeLine(limb, orientation, columnNum):
        lines = []
        times = []
        currMin = 0
        for i in range(len(strLines)):
            if i < len(strLines)-2:#last two lines are for time data
                lines.append(eval(strLines[i]))
            else:
                str2 = strLines[i].split(': ')
                times.append(float(str2[1]))
        prevTime = lines[0][TIME]
        AGG_MOVEMENT_DURING_INTERVAL = 0 
        for i in range(len(strLines)-4):#does -4 so it doesnt go out of bounds and hit information about the fire
            if (currMin < 10):
                if lines[i][TIME] < times[FIRE_TIME]:#ignore everything before fire comes on
                    continue
                if lines[i][TIME] > times[FIRE_TIME] + 3:#ignore everything after interval
                    continue
                print limb
                print lines[i][limb]
                print lines[i+1][limb]
                y = [0, dist(lines[i][limb], lines[(i + 1)][limb], orientation)]
                print "aggregate"
                AGG_MOVEMENT_DURING_INTERVAL += y[1]
                print AGG_MOVEMENT_DURING_INTERVAL
                if lines[i][TIME] > (prevTime + 60):
                    currMin += 1
                    prevTime = prevTime + 60 
        putOnExcel(fileName.split('.')[0], AGG_MOVEMENT_DURING_INTERVAL, columnNum)  

#determine which row participant is in in the table
def determineRowForParticipant(participantName):
    workbook = load_workbook('/Users/angelolvera/Desktop/ThirdArmData.xlsx')
    ws = workbook.get_sheet_by_name('Sheet1')
    rowNum = 2
    while(True):
        #print ws.cell('A1').value
        #print ws.cell('%s%s'%('A',rowNum)).value
        if(ws.cell('%s%s'%('A',rowNum)).value == participantName):
            #print "not equal to participantName"
           # print "1"
            print ws.cell('%s%s'%('A',rowNum)).value
            return rowNum
        elif(ws.cell('%s%s'%('A',rowNum)).value == None):
           # print "2"
            ws.cell('%s%s'%('A',rowNum)).value = participantName
            print ws.cell('%s%s'%('A',rowNum)).value
            return rowNum
        #print "3"
        print ws.cell('%s%s'%('A',rowNum)).value
        rowNum += 1

#determine row for participant name-> use first free row if not there
#print movement data in correct column given by columnNum
def putOnExcel(participantName,  movement, columnNum):
    #workbook = load_workbook('C:\Users\\angelolvera\Desktop\ThirdArmData.xlsx')
    workbook = load_workbook('C:\Users\Abhishek\Google Drive\Desktop\Analyzing Data\ThirdArmData.xlsx')
    ws = workbook.get_sheet_by_name('Sheet1')
    #print ws
    ws.cell('%s%s'%(get_column_letter(columnNum),determineRowForParticipant(participantName))).value = movement
    #print "limbName: " + LIMB_NAMES[limbName] + " orientation: " + ORIENTATIONS[orientation] + " value: " + str(movement)
    workbook.save('/Users/angelolvera/Desktop/ThirdArmData.xlsx')


#////////////////////////    
if not os.path.exists(dirPre + "NewDataPlotsAngel"):
        os.mkdir(dirPre + "NewDataPlotsAngel")
rowInExcelFile = 0
for fileName in dirList:
        if fileName.endswith(".dat"):
                print "Plotting " + fileName + "..."
                file = open(fileName, 'r')
                strLines = file.readlines()
                
                if len(strLines) == 0:
                        print "File " + fileName + " empty, on to next!"
                        continue 
                columnNum = 0
                makeLine(HEAD_PPT_EULER,YAW, columnNum = 2)
                makeLine(LH_PPT_EULER,PITCH, columnNum =  3)
                makeLine(HEAD_PPT_EULER,ROLL, columnNum = 4)
                makeLine(HEAD_PPT_POS,X_POS, columnNum = 5)
                makeLine(HEAD_PPT_POS,Y_POS, columnNum = 6)
                makeLine(HEAD_PPT_POS,Z_POS, columnNum = 7)
                makeLine(LH_PPT_EULER,YAW, columnNum = 8)
                makeLine(LH_PPT_EULER,PITCH, columnNum =  9)
                makeLine(LH_PPT_EULER,ROLL, columnNum = 10)
                makeLine(LH, X_POS, columnNum = 11)
                makeLine(LH, Y_POS, columnNum = 12)
                makeLine(LH, Z_POS, columnNum = 13)
                makeLine(RH_PPT_EULER,YAW, columnNum = 14)
                makeLine(RH_PPT_EULER,PITCH, columnNum =  15)
                makeLine(RH_PPT_EULER,ROLL, columnNum = 16)
                makeLine(RH, X_POS, columnNum = 17)
                makeLine(RH, Y_POS, columnNum = 18)
                makeLine(RH, Z_POS, columnNum = 19)                
                print "Finished!"
