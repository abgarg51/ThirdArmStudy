import os
from matplotlib import pyplot
import math
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook#for writing to excel file

#Indexes for data
LH = 0
RH = 1
HEAD_PPT_POS = 2
HEAD_PPT_EULER = 3
LH_PPT_EULER = 4
RH_PPT_EULER = 5
DATE = 6
TIME = 7

YAW = 1
PITCH = 2
ROLL = 3

TARGET_TIME = 0
FIRE_TIME = 1
INTERVAL = 120
LIMB_NAMES = ['LeftHand', 'RightHand', 'HeadPosition', 'HeadEuler', 'LeftHandEuler', 'RightHandEuler']
NUM_MIN = 9
ORIENTATIONS = ['', 'YAW' ,'PITCH', 'ROLL']#0 means orientation not applicable
dirList = os.listdir(os.path.curdir)
dirPre = ""
#FUNCTIONS
def dist(p1, p2, orientation):
    if orientation == 0:
        return math.sqrt(math.pow(p2[0] - p1[0], 2) + math.pow(p2[1] - p1[1], 2) + math.pow(p2[2] - p1[2],2))#squarerrot((x1-x2)^2+(y1-y2)^2+(z1-z2)^2)
    else:
        distanceCalculated = abs(p2[orientation - 1] - p1[orientation - 1])
        if distanceCalculated > 180:
            return 360 - distanceCalculated
        else:
            return distanceCalculated
#uses a different distance function to calculate changes in position for the intersense

#outputs a graph to the file with the given limb and orientation movement graph
def makeLine(limb, orientation):
        lines = []
        timelines = []
        times = []
        TARGET_FRAME = 0
        FIRE_FRAME = 0
        pyplot.clf()
        currMin = 0
        for i in range(len(strLines)):
            if i < len(strLines)-2:
                lines.append(eval(strLines[i]))
            else:
                timelines.append(strLines[i])    
        for i in range(len(timelines)):
            str = timelines[i]
            str2 = str.split(': ')
            times.append(float(str2[1]))
        prevTime = lines[0][TIME]
        AGG_MOVEMENT_DURING_INTERVAL = 0 
        for i in range(len(strLines)-4):#does -4 so it doesnt go out of bounds and hit information about the fire
            if (currMin < 10):
                y = [0, dist(lines[i][limb], lines[(i + 1)][limb], orientation)]
                x = [i, i]
                pyplot.plot(x,y)
                if orientation == 0:
                    v = [0, 15000, 0, .05]
                else:
                    v = [0, 15000, 0, 20]
                pyplot.axis(v)
                if lines[i][TIME] < times[TARGET_TIME]:
                    TARGET_FRAME = i
                if lines[i][TIME] < times[FIRE_TIME]:
                    FIRE_FRAME = i
                if lines[i][TIME] < times[TARGET_TIME] and lines[i][TIME] < times[FIRE_TIME]:#is inside interval
                    AGG_MOVEMENT_DURING_INTERVAL += y[1]
                if lines[i][TIME] > (prevTime + 60):
                    currMin += 1
                    prevTime = prevTime + 60
        pyplot.axvline(x=TARGET_FRAME, color='b')
        pyplot.axvline(x=FIRE_FRAME, color='r')
        pyplot.axvline(x=FIRE_FRAME + INTERVAL, color='r')
        pyplot.savefig(dirPre + "NewDataPlotsAngel/" + fileName.split('.')[0] + "_As_Spikes" + LIMB_NAMES[limb] + "_" + 
            ORIENTATIONS[orientation] + "_Plot.png")  
        putOnExcel(fileName.split('.')[0], limb, orientation, AGG_MOVEMENT_DURING_INTERVAL)  
def calculateColumn(limbName, orientation):
    return get_column_letter((limbName + 2) + orientation * 2)


def putOnExcel(participantName, limbName, orientation, movement):
    workbook = load_workbook('C:\Users\Abhishek\Google Drive\Desktop\Analyzing Data\NewDataPlotsAngel\ThirdArmData.xlsx')
    ws = workbook.get_sheet_by_name('Sheet1')
    rowNum = 1
    print ws
    while(True):
        #print ws.cell('A1').value
        print ws.cell('%s%s'%('A',rowNum)).value
        if(ws.cell('%s%s'%('A',rowNum)).value == participantName):
            print "not equal to participantName"
            return
        elif(ws.cell('%s%s'%('A',rowNum)).value == None):
            print participantName
            ws.cell('%s%s'%('A',rowNum)).value = participantName
            break
        rowNum += 1
    ws.cell('%s%s'%(calculateColumn(limbName, orientation),rowNum)).value = movement
    print "limbName: " + LIMB_NAMES[limbName] + " orientation: " + ORIENTATIONS[orientation] + " value: " + str(movement)
    workbook.save('C:\Users\Abhishek\Google Drive\Desktop\Analyzing Data\NewDataPlotsAngel\ThirdArmData.xlsx')





#////////////////////////    
if not os.path.exists(dirPre + "NewDataPlotsAngel"):
        os.mkdir(dirPre + "NewDataPlotsAngel")
rowInExcelFile = 0
for fileName in dirList:
        if fileName.endswith(".dat"):
                print "Plotting " + fileName + "..."
                file = open(fileName, 'r')
                strLines = file.readlines()
                
                if len(strLines) == 0:
                        print "File " + fileName + " empty, on to next!"
                        continue 
                makeLine(RH,0)
                makeLine(LH,0)
                makeLine(LH_PPT_EULER,YAW)
                makeLine(RH_PPT_EULER,YAW)
                makeLine(LH_PPT_EULER,PITCH)
                makeLine(RH_PPT_EULER,PITCH)
                makeLine(LH_PPT_EULER,ROLL)
                makeLine(RH_PPT_EULER,ROLL)
                
                print "Finished!"

